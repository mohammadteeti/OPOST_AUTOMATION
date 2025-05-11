
import tkinter as tk
from tkinter import  filedialog , messagebox    
import sys
import threading
import os
import time
import subprocess
import webbrowser
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.styles import Border, Side
import time
from datetime import datetime,timedelta

from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService

from selenium.webdriver.edge.service import Service as EdgeService
from webdriver_manager.core.manager import DriverManager


from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import re 
import psutil
import random 
import winsound
import shutil

red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
green_fill =PatternFill(start_color="00FF00",end_color="00FF00",fill_type="solid")

employee_urls = []
debugging_mode_string =""


time_difference_per_user = [] 
shipment_numbers= []
pattern = re.compile(r'^(0[1-9]|1[0-2])-(0[1-9]|[12][0-9]|3[01])$')
shift_time_T=0
random_sample=0
stop_event =  threading.Event()


name=""

browser_version=""





#End the Edge webview of each run 
def kill_process_by_name(process_name):
    for process in psutil.process_iter(['pid', 'name']):
        if process.info['name'] and process_name.lower() in process.info['name'].lower():
            try:
                psutil.Process(process.info['pid']).terminate()
                print(f"Killed process: {process.info['name']} (PID: {process.info['pid']})")
            except psutil.NoSuchProcess as e1:
                print(f"error in kill_debugging_edge {e1} ")
                light_label.config(background="#FF0000")
            except psutil.AccessDenied:
                print(f"Access denied for process: {process.info['name']} (PID: {process.info['pid']})")
                light_label.config(background="#FF0000")
            except Exception as e:
                print(f"Error terminating process {process.info['name']}: {e}")
                light_label.config(background="#FF0000")


def kill_debugging_chrome():
    for process in psutil.process_iter(attrs=['pid', 'name', 'cmdline']):
        try:
            if "chrome.exe" in process.info['name'].lower():
                cmdline = " ".join(process.info['cmdline']) if process.info['cmdline'] else ""
                
                # Check if Chrome was launched with remote debugging port 9333
                if "--remote-debugging-port=9333" in cmdline:
                    print(f"Killing debugging Chrome process: {process.info['name']} (PID: {process.info['pid']})")
                    psutil.Process(process.info['pid']).terminate()
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess) as e:
            print(f"error in kill_debugging_edge {e} ")
            light_label.config(background="#FF0000")




def kill_debugging_edge():
    for process in psutil.process_iter(attrs=['pid', 'name', 'cmdline']):
        try:
            if "msedge.exe" in process.info['name'].lower():
                cmdline = " ".join(process.info['cmdline']) if process.info['cmdline'] else ""
                
                # Check if Edge was launched with remote debugging port 9333
                if "--remote-debugging-port=9333" in cmdline:
                    print(f"Killing Edge debugging process: {process.info['name']} (PID: {process.info['pid']})")
                    psutil.Process(process.info['pid']).terminate()
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess)as e:
            light_label.config(background="#FF0000")
            print(f"error in kill_debugging_edge {e} ")



kill_debugging_chrome()
kill_debugging_edge()
kill_process_by_name("msedgewebview2.exe")


#read the preveious versions of the browser , in order to skip version manager if the same version is used 
with open("chrome_version.dat","r") as f :
        global prev_chrome_version
        prev_chrome_version=f.read().strip()
        f.close

with open("edge_version.dat","r") as f :
        global prev_edge_version
        prev_edge_version=f.read().strip()
        f.close


def get_browser_version():

    cmd="version.bat"
    result=subprocess.run(cmd,shell=True,capture_output=True)
    chrome_version=""
    edge_version=""
        # Define the WMIC command

    time.sleep(4)

    with open("chrome_version.dat" , "r") as f:
        chrome_version=f.read().strip()
        f.close()
    
    with open("edge_version.dat" , "r") as f:
        edge_version=f.read().strip()
        f.close()

    return [chrome_version,edge_version]

####################################################### REDIRECT OUTPUT CLASS ###############################################################
class RedirectOutput : 

    def __init__(self,text_widget):
        self.text_widget=text_widget

    def write(self,text):
        self.text_widget.insert(tk.END,text)
        self.text_widget.see (tk.END)

    def flush(self):
        pass
#############################################################################################################################################




######################################################### START BROWSER SESSIONS CODE ########################################################



def start_chrome_session(browser_version,debugging_string):
    chrome_path =str("C:\\Program Files\\Google\\Chrome\\Application")
    os.environ["PATH"] +=os.pathsep+ chrome_path
    
    time.sleep(0.5)
    
    # Ensure the debugging command is properly defined
    cmd = debugging_string  # Example: "chrome.exe --remote-debugging-port=9333 --user-data-dir='C:\\selenium\\'"

    if not cmd:
        print("Debugging mode command is empty. Check the configuration.\n")
        return

    try:
        # Use subprocess.Popen to start Chrome in the background
        process = subprocess.Popen(cmd, shell=True)
        print(f"Chrome session started with PID: {process.pid}\n")
        msg= messagebox.askyesno("Login!","Is your login to OPOST account ready ? ")
        while( not msg):
            msg= messagebox.askyesno("Login!","Is your login to OPOST account ready ? ")

        options=webdriver.ChromeOptions()
        options.debugger_address = "127.0.0.1:"+str(port).strip()
        options.add_argument("--headless=new")  # Run Chrome in headless mode to avoide 	GetHandleVerifier [0x00007FF6FFA00AF5+13637] error during the normal usage of system by user 
        
        if prev_chrome_version == browser_version :
            print("Using Same Driver Version")
            
            driver=webdriver.Chrome(options=options)
        else :
            # Initialize WebDriver with the correct ChromeDriver version
            print("Installing New Driver Version")
            driver = webdriver.Chrome(
            service=ChromeService(ChromeDriverManager(browser_version).install()),
            options=options
            )

        
        print("webdriver instatiated correctly")
        light_label.config(background="#00FF00")
        get_employee_data_from_excel(input_path,driver)
    except Exception as e:
        print(f"Error starting Chrome session: {e}\n")
        light_label.config(background="#FF0000")


def start_edge_session(browser_version,debugging_string):
    
    edge_path=str("c:\\program files (x86)\\Microsoft\\Edge\\Application")

    os.environ["PATH"]+=os.pathsep + edge_path
    time.sleep(0.5)
    # Ensure the debugging command is properly defined
    cmd = debugging_string  # Example: "chrome.exe --remote-debugging-port=9333 --user-data-dir='C:\\selenium\\'"
    
    if not cmd:
        print("Debugging mode command is empty. Check the configuration.\n")
        return

    try:
        # Use subprocess.Popen to start Chrome in the background
        process = subprocess.Popen(cmd, shell=True)
        print(f"Edge session started with PID: {process.pid}\n")
        msg= messagebox.askyesno("Login!","Is your login to OPOST account ready ? ")
        while( not msg):
            msg= messagebox.askyesno("Login!","Is your login to OPOST account ready ? ")
            
        options=webdriver.EdgeOptions()
        options.debugger_address = "127.0.0.1:"+str(port).strip()
        options.add_argument("--headless=new")  # Run Chrome in headless mode to avoide 	GetHandleVerifier [0x00007FF6FFA00AF5+13637] error during the normal usage of system by user 

        if prev_edge_version == browser_version :
            print("Using Same Driver Version")
            driver=webdriver.Edge(options=options)
        else :
            # Initialize the WebDriver with the correct EdgeDriver version
            print("Installing New Driver Version")
            driver = webdriver.Edge(
            service=EdgeService(EdgeChromiumDriverManager(browser_version).install()),
            options=options
            )
        
        print("webdriver instatiated correctly")
        light_label.config(background="#00FF00") 
        get_employee_data_from_excel(input_path,driver)
    except Exception as e:
        print(f"Error starting Chrome session: {e}\n")
        light_label.config(background="#FF0000")
##########################################################################################################################################


def browsFile(): 
    global input_path
    input_path = filedialog.askopenfilenames(filetypes=[("Excel Files","*.xlsx"),("All Files","*.*")])

    if input_path:
        path_box.delete(0,tk.END)
        path_box.insert(0,input_path)

        
        

def run_script ():
    global browserChoice
    browserChoice= browser_var.get()
    #input_path =path_box.get()
    


    if not browserChoice:
        messagebox.showerror("Error","Please Select Browser Type!")
        return
    
    if not input_path : 
        messagebox.showerror("Error","Please Enter Employee Text File Path")
        return
    
    if not entry_username.get():
        messagebox.showerror("Error","Please Enter UserName/Email of your  Optimus Account")
        return
    if not entry_password.get():
        messagebox.showerror("Error","Please Enter Password of your  Optimus Account")
        return
    
    def main_processing_code():
        
        try :
            print('Starting Program .... \n')
            print("Reading configuration file ...\n")
            with open("config.cfg","r",encoding="UTF-8") as cfg:
                global port
                debugging_mode_string= cfg.readline().split(",") #the first line in Config file contains the whole debugging string from which chrome and edge debugging mode are extracted
                port=cfg.readline().split(",")[1]
                cfg.close()

            browser_version=get_browser_version()
            if  browserChoice.lower()=="chrome":
                try:
                    start_chrome_session(browser_version=browser_version[0],debugging_string=debugging_mode_string[1])
                except Exception as e:
                    print(f"ُmain_processing_code Function calling start_chrome_session function : \n  {e}")
                    light_label.config(background="#FF0000")
            elif browserChoice.lower()=="edge":
                try:               
                    start_edge_session(browser_version=browser_version[1],debugging_string=debugging_mode_string[2])
                except Exception as e :
                    print(f"ُmain_processing_code Function calling start_edge_session function \n  {e}")
                    light_label.config(background="#FF0000")

        except Exception as e: 
            messagebox.showerror("Error" , e)
            light_label.config(background="#FF0000")
            
    
    global thread
    thread = threading.Thread(target=main_processing_code,daemon=True)
    stop_event.clear()
    thread.start()


def stop_main_processing_thread():
    stop_event.set()
    kill_process_by_name("msedgewebview2.exe")
    kill_debugging_chrome()
    kill_debugging_edge()
    light_label.config(background="#FF0000")
    

def show_about_window():
    about_win = tk.Toplevel(root)
    about_win.title("About")
    about_win.geometry("300x200")
    
    # Name label
    name_label = tk.Label(about_win, text="Created By:\nEng. Mohammad Jbber Teeti", font=("Arial", 12, "bold"))
    name_label.pack(pady=(10, 5))

    # GitHub button
    github_button = tk.Button(
        about_win, 
        text="Visit GitHub", 
        command=lambda: webbrowser.open("https://github.com/mohammadteeti")
    )
    github_button.pack(pady=5)

    # LinkedIn button
    linkedin_button = tk.Button(
        about_win, 
        text="Visit LinkedIn", 
        command=lambda: webbrowser.open("https://www.linkedin.com/in/mohammadteeti/")
    )
    linkedin_button.pack(pady=5)

    # Close button
    close_button = tk.Button(about_win, text="Close", command=about_win.destroy)
    close_button.pack(pady=10)


#####################################################################################################################################
#####################################################################################################################################
################################################### The major code processing #######################################################

def get_employee_data_from_excel(input_path,driver):
    

    browserChoice=browser_var.get()
    print (f"\nChoice : {browserChoice.lower()}")

#try to login first to avoid selenium crash   exception in reading tracking number 
    try :
            driver.get("https://opost.ps/login")
            time.sleep(1)
            email_field=driver.find_element(By.ID,"email")
            password_field = driver.find_element(By.ID,"password")

            time.sleep(1)
            # Use JavaScript to clear autofilled values
            driver.execute_script("document.getElementById('email').value = '';")
            driver.execute_script("document.getElementById('password').value = '';")

            
            time.sleep(2)

            email_field.send_keys(entry_username.get())
            password_field.send_keys(entry_password.get())

            driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
            print("Login submitted!") 
    except NoSuchElementException :
            #if email element is not found , then most probably the user is already logged in 
            print(f"\nMay be you are logged in already !") 
            light_label.config(background="#FFFF00")
            




    for file  in input_path: #skip first row (Headers)
        if stop_event.is_set():
                break
            
        name =os.path.basename(file).split(".")[0].split(" ")[1].strip() #get the name of the file without the extension
        path=file
        file_date =os.path.basename(file).split(".")[0].split(" ")[0].strip() #get the date from the file name
        
        print (f'name : {name} , date : {file_date} , path : {path} , is_random : {is_random}')
        
        time_difference_per_user = [] 
        cod_count=0
        shipment_numbers=[]
        reply_times = []
        pending_type= ""
        # Open the provided Excel file and read the B column from row 2 onward in the first sheet
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        tracking_numbers = []
        for cell in ws['B'][1:]:  # Skip the first row (header)
            if cell.value:
                tracking_numbers.append(cell.value)

        if is_random==1:
                print(f"Random of {random_sample} Samples are Chosen\n")
                tracking_numbers= get_random_tracking_numbers(tracking_numbers) 
        else:
                print("Full File is Chosen\n")
        
        for number in tracking_numbers:
            if stop_event.is_set():
                break
            print(f"{tracking_numbers.index(number)}: Working On {name} with Number : {number} in the Date : {file_date}\n")
            try:
                driver.execute_script(f"window.open('https://opost.ps/resources/shipments?tracking_number={number}', '_self');")
            except Exception as e:
                print(f" Error in get_employee_data_from_excel Function : \n  {e}")
                light_label.config(background="#FF0000")

            # Switch to the new tab
            driver.switch_to.window(driver.window_handles[-1])

            # Optionally, wait for the page to fully load
            time.sleep(3)

            # Find the 29th button on the page and click it
            try :
                buttons = driver.find_elements(By.TAG_NAME, "button")
                if len(buttons) > 28:
                    button = buttons[28]
                    button.click()
                else:
                    print("Button not found\nTrying Again One Time... ")
                    time.sleep(1)
                    try :
                        buttons=driver.find_elements(By.TAG_NAME, "button")
                        if len(buttons)>28:
                            button=buttons[28]
                            button.click()
                        else:
                            print("Button not found\nPlease Check Internet Connection and try again ")
                            winsound.Beep(600,1000) 
                            continue
                    except Exception as e:
                        print(f"{e}")
                        continue
            except Exception as e :
                print(f"{e}")
                continue
            light_label.config(background="#00FF00")
            # Wait for the new content to load
            try:
                # Define the locator for the new content
                new_content_locator = (By.XPATH, "//button[contains(., 'notes')]")  # the notes text is the only unique modifier  in the pop up opost dialog , it's text is never taken along the html flow of the page
                
                #this allocation is used only to "Take Time " , we give the system few time to make sure the content is fully loaded 
                new_content = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located(new_content_locator)
                )
                
                #The following allocation is the real content allocation , from the loaded content below we extract data
                new_content_locator = (By.CSS_SELECTOR, "tr")  # since the dialog is a table of rows , the main elemetn to target is the table rows collection
                new_content = WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located(new_content_locator)
                )
                
                # Once the new Table  is loaded, we can now parse the table or other new elements
                table_row = driver.find_elements(*new_content_locator)  #   retrieves  the collection of all rows 
                first_pending_of_driver =""                             #   define a variable to carry the first pending status 
                first_pending_of_employee=""                            #   define a variable to carry the first employee pending just after the driver firtst pending 
                
                
                is_first_time_employee_pen_detected=False                            #   since we catch all the dialog rows , we are to count the number of COD for each number in a defined date, 
                                                                        #   the COD status may repeate as well as the Pending status , we are targetting the first pending only 
                                                                        #   so the {is_first_time_detected} is put to control that logic
                is_first_time_driver_pen_detected=False
                # Table rows collection  is firstly reversed , to start from the 'Submitted'  status and forth on   
                # The logic below consists the results to the Pending and COD status types only, other status types are eventually ignored                                           
                # there is a bug here where the readign start from above

                
                for row in reversed(table_row):
                    if "COD Pickup" in row.text and file_date in row.text: # add the date to the condition to avoid the COD status that is not related to the current date
                        cod_count =cod_count+1
                        break

                        
                
                for row in reversed(table_row):

                    
                    if  "Pending" in row.text and file_date in row.text:
                        td_elements = row.find_elements(By.CSS_SELECTOR, "td")


                        # Iterate through td elements and print their text
                        pending_data = [td.text for td in td_elements]
                        print(f"{pending_data[1]} {pending_data[3]}\n")
                        #print (f"Pending Type  :  {pending_data[5]}")
                        
                        if not is_first_time_driver_pen_detected:
                            if pending_data[1].strip() == pending_data[3].strip():
                                first_pending_of_driver=pending_data[0]
                                pending_type=pending_data[5]
                                is_first_time_driver_pen_detected=True
                            
                        if not is_first_time_employee_pen_detected :
                            if is_first_time_driver_pen_detected:
                                if name in pending_data[1] :# check_if_name_occures_in_pending_line(pending_data[1]): #'291لارا' in  pending_data[1]  or '296هبة' in  pending_data[1] or '290رند' in  pending_data[1] or '294حمزة' in  pending_data[1] or 'احمد295' in  pending_data[1] or 'متابعة عوالق' in  pending_data[1] :
                                    first_pending_of_employee=modify_time_if_before_T(pending_data[0])
                                    
                                    is_first_time_employee_pen_detected=True
                                #break
                    if is_first_time_employee_pen_detected: # no need to read the rest or row as the first pending variables are assigned
                        pending_data=[]
                        break

                        
                        
                    pending_data=[]
                       

                if first_pending_of_employee and first_pending_of_driver:
                    #print(f"time of employee {first_pending_of_employee} \ntime of driver {first_pending_of_driver}\n") #   show Pending resluts for both employee and driver
                    time1 = datetime.strptime(first_pending_of_driver, "%Y-%m-%d %H:%M:%S").time()                      #   extract the time only from the full date-time format of driver Pending status 
                    time2 = datetime.strptime(first_pending_of_employee, "%Y-%m-%d %H:%M:%S").time()                    #   extract the time only from the full date-time format of employee Pending status 
                else:
                    time1=datetime.strptime("2000-01-01 00:00:00", "%Y-%m-%d %H:%M:%S").time() 
                    time2=datetime.strptime("2000-01-01 00:00:00", "%Y-%m-%d %H:%M:%S").time() 
                    
                
                # Convert the time components to timedelta objects 
                time1_delta = timedelta(hours=time1.hour, minutes=time1.minute, seconds=time1.second)
                time2_delta = timedelta(hours=time2.hour, minutes=time2.minute, seconds=time2.second)

                #print(f"Timedelta1 {time1_delta}\n")
                #print(f"Timedelta2 {time2_delta}\n")
                # Calculate the absolute difference in minutes between the two time components
                time_difference = abs(time2_delta - time1_delta)
                difference_in_minutes = time_difference.total_seconds() / 60

                # Print the difference in minutes
                print(f"The difference in minutes is: {difference_in_minutes:.2f}\n")
                time_difference_per_user.append(round(difference_in_minutes,2))
            
                shipment_numbers.append(number)
                reply_times.append((first_pending_of_employee,first_pending_of_driver,pending_type)) #store the response times of both employee and driver for each number

            
            except Exception as e:
                    print("New content did not load within the wait time:", e,"\n")
                    winsound.Beep(700,1000)
                    continue #  The flow should continue and ignore any exceptions as the exceptions are mainly generated pair Tracking Number 
            
             
        
        #call function to create the results as excel file 
        
        
        wb.close()
        create_excel(file_date, time_difference_per_user,cod_count,shipment_numbers,reply_times,name)


############ Auxilary Function #################

def get_random_tracking_numbers(tracking_numbers_list):
    
    # Ensure the original list has at least 20 elements
    if len(tracking_numbers_list) >= random_sample:
        random_numbers = random.sample(tracking_numbers_list, random_sample)
        return random_numbers
    else:
        print("The list does not contain enough elements. Returning the whole tracking numbers list \n")
        return  tracking_numbers_list
    

def modify_time_if_before_T(datetime_str):
    """
    Modify the time part of the datetime string to 10:00:00 if the hour is before T:00:00 am.
    
    Parameters:
    datetime_str (str): The input datetime string in the format 'YYYY-MM-DD HH:MM:SS'.
    
    Returns:
    str: The modified datetime string.
    """
    # Parse the datetime string into a datetime object
    dt = datetime.strptime(datetime_str, '%Y-%m-%d %H:%M:%S')
    
    # Check if the hour is before 10
    if dt.hour < shift_time_T:
        # Modify the time to 10:00:00
        dt = dt.replace(hour=shift_time_T, minute=0, second=0)
    
    # Convert the datetime object back to a string
    modified_datetime_str = dt.strftime('%Y-%m-%d %H:%M:%S')
    
    return modified_datetime_str


    
def create_excel(date, employee_data,cod_count,shipment_numbers,reply_times, user_name):
    try:
        
        # Define a thin solid border style
        thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
)

        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Set the date in cell A1
        ws['A1'] = date

        # Write the header for the single user
        ws['B1'] = user_name

        # Write the header for the Shipment Numbers
        ws['C1'] = "رقم الشحنة"
        
        #write the header for the COD count
        ws['D1'] = "في حال تم التسليم بدون عالق"
        
        #write the header for the reply times
        ws['E1'] = "وقت الرد من الموظف"
        ws['F1'] = "وقت الرد من السائق"
        ws['G1'] = "نوع العالق"
        
        ws["A1"].font = openpyxl.styles.Font(bold=True)
        ws["B1"].font = openpyxl.styles.Font(bold=True)
        ws["C1"].font = openpyxl.styles.Font(bold=True)
        ws["D1"].font = openpyxl.styles.Font(bold=True)
        ws["E1"].font = openpyxl.styles.Font(bold=True)
        ws["F1"].font = openpyxl.styles.Font(bold=True)
        ws["G1"].font = openpyxl.styles.Font(bold=True)
        
        ws["A1"].border = thin_border
        ws["B1"].border = thin_border
        ws["C1"].border = thin_border
        ws["D1"].border = thin_border
        ws["E1"].border = thin_border
        ws["F1"].border = thin_border
        ws["G1"].border = thin_border
        
        #set columns widths 
        ws.column_dimensions["A"].width=len("COD COUNT = ")
        ws.column_dimensions["B"].width=len(str(user_name)+" ")
        ws.column_dimensions["C"].width=len("XXXXXX-XXX-XXXXXXXXX  ")
        ws.column_dimensions["D"].width=len(" تم التسليم بدون عالق - لا تحسب في التقرير")


        #define a varianle to count shipments that have been COD without any pending status
        cod_without_pending=0
        # Write the data for the single user
        for row_num, value in enumerate(employee_data, start=2):
            ws.cell(row=row_num, column=2).value = value
            if value>10 :
                ws.cell(row=row_num, column=2).fill=red_fill

            if value == 0 :
                ws.cell(row=row_num,column=4).value="تم التسليم بدون عالق - لا تحسب في التقرير"
                ws.cell(row=row_num,column=4).fill=green_fill
                ws.cell(row=row_num,column=4).border = thin_border
                cod_without_pending=cod_without_pending+1
            ws.cell(row=row_num,column=2).border = thin_border
            
        for row_num  ,value in enumerate(shipment_numbers,start=2):
            ws.cell(row=row_num,column=3).value=value
            ws.cell(row=row_num,column=3).border = thin_border
            
        for row_num ,value in enumerate(reply_times,start=2):
            ws.cell(row=row_num,column=5).value=value[0]
            ws.cell(row=row_num,column=6).value=value[1]
            ws.cell(row=row_num,column=7).value=value[2]
            
            ws.cell(row=row_num,column=5).border = thin_border
            ws.cell(row=row_num,column=6).border = thin_border
            ws.cell(row=row_num,column=7).border = thin_border
            
        if len(employee_data) == 0 :
            employee_data=[1]
        ws.cell(row=len(employee_data)+3,column=1 ).value="Average = " 
        ws.cell(row=len(employee_data)+3,column=1 ).border = thin_border
        try:
            ws.cell(row=len(employee_data)+3,column=2 ).value= round(sum(employee_data)/(len(employee_data)-cod_without_pending),2)
            ws.cell(row=len(employee_data)+3,column=2 ).border = thin_border
        except ZeroDivisionError:
            ws.cell(row=len(employee_data)+3,column=2 ).value=0
            ws.cell(row=len(employee_data)+3,column=2 ).fill=red_fill
            ws.cell(row=len(employee_data)+3,column=2 ).border = thin_border
            
        ws.cell(row=len(employee_data)+5,column=1 ).value="COD COUNT = " 
        ws.cell(row=len(employee_data)+5,column=2 ).value=cod_count
        ws.cell(row=len(employee_data)+5,column=2 ).border = thin_border
        ws.cell(row=len(employee_data)+5,column=1 ).border = thin_border


        # Save the workbook
        file_name = f"{date.replace('/', '-')}_for_{user_name}.xlsx"
        wb.save(file_name)
        winsound.Beep(900,200)
        time.sleep(0.2)
        winsound.Beep(900,200)
        time.sleep(0.2)
        winsound.Beep(900,200)
        print(f"Excel file '{file_name}' created successfully.\n")

    except   Exception as e : 
        print(f"Error in Creating Excel Function {e}\n")



def update_shift_time(value):
    global shift_time_T
    shift_time_T=int(value)

def update_sample_number(value):
    global random_sample
    random_sample=value

def update_is_random_list(value):
    global is_random
    is_random=value == "True"

#####################################################################################################################################
#####################################################################################################################################


##################################### Create Main Screen Window with widgets ########################################################



# Create the main application window
root = tk.Tk()
root.title("Pending Response Monitor")
root.geometry("600x600")
root.resizable(False, False)
root.config(bg="#cceeff")

# Create a frame for the file selection
file_frame = tk.Frame(root, padx=10, pady=10, bd=2, relief=tk.SOLID)
file_frame.pack(fill=tk.X, padx=10, pady=10)

# Browse button and path display
browse_button = tk.Button(file_frame, text="Brows Button", command=browsFile, width=12)
browse_button.pack(side=tk.LEFT, padx=(0, 5))

path_box = tk.Entry(file_frame, width=50)
path_box.pack(side=tk.LEFT, fill=tk.X, expand=True)

# Create a frame for browser selection
browser_frame = tk.Frame(root, padx=10, pady=10)
browser_frame.pack(fill=tk.X)
browser_frame.config(bg="#cceeff")
# create red- and green 
light_label=tk.Label(file_frame,text="",width=3,height=3,background="#FF0000")
light_label.pack(padx=20)
# Create a Tkinter variable
browser_var = tk.StringVar(value="xxx")  # Set to a value that doesn't match any button

chrome_radio = tk.Radiobutton(browser_frame, text="Chrome", variable=browser_var, value="Chrome")
chrome_radio.pack(side=tk.LEFT, padx=10)

edge_radio = tk.Radiobutton(browser_frame, text="Edge", variable=browser_var, value="Edge")
edge_radio.pack(side=tk.LEFT, padx=10)



# Create a frame for shift time selection
shift_time_frame = tk.Frame(root, padx=10, pady=10, bg="#cceeff")
shift_time_frame.pack(fill=tk.X)

# Label and dropdown for shift time
shift_time_label = tk.Label(shift_time_frame, text="Shift Time Start (T):", bg="#cceeff")
shift_time_label.pack(side=tk.LEFT, padx=5)

shift_time_var = tk.IntVar(value=9)
shift_time_menu = tk.OptionMenu(shift_time_frame, shift_time_var, 9, 10, command=update_shift_time)
shift_time_menu.pack(side=tk.LEFT, padx=5)

# Label and dropdown for is_random_sample
is_random_sample_label = tk.Label(shift_time_frame, text="Take Random Sample", bg="#cceeff")
is_random_sample_label.pack(side=tk.LEFT, padx=5)

is_random_sample = tk.StringVar(value="False")
is_random_sample_menu = tk.OptionMenu(shift_time_frame, is_random_sample, "True", "False", command=update_is_random_list)
is_random_sample_menu.pack(side=tk.LEFT, padx=5)

# Label and dropdown for random sample size
random_selector_label = tk.Label(shift_time_frame, text="Sample Size:", bg="#cceeff")
random_selector_label.pack(side=tk.LEFT, padx=5)

default_random_sample = tk.IntVar(value=5)
random_selector_menu = tk.OptionMenu(shift_time_frame, default_random_sample, 5, 10, 15, 20, 25, 30, command=update_sample_number)
random_selector_menu.pack(side=tk.LEFT, padx=5)



#create login frame and controls 

# Create login login_frame
login_frame = tk.Frame(root, padx=10, pady=2)
login_frame.pack(pady=20)
login_frame.config(bg="#cceeff")
# Username label and entry
label_username = tk.Label(login_frame, text="Username:")
label_username.grid(row=0, column=0, sticky="w")
entry_username = tk.Entry(login_frame,width=40)
entry_username.grid(row=0, column=1)

# Password label and entry
label_password = tk.Label(login_frame, text="Password:")
label_password.grid(row=1, column=0, sticky="w")
entry_password = tk.Entry(login_frame, show="*",width=40)
entry_password.grid(row=1, column=1)

# Login button 
#the login process is done in the run button
#button_login = tk.Button(login_frame, text="Login")
#button_login.grid(row=2, column=0, columnspan=2, pady=10)

# Create a button to run the script
run_button = tk.Button(root, text="Run", command=run_script ,font=("Arial", 12, "bold"))
run_button.pack(pady=5,padx=5)

#create a button to stop the script
stop_button = tk.Button(root,text="Stop",command=stop_main_processing_thread ,font=("Arial", 12, "bold"))
stop_button.pack(padx=10 ,pady=10)

#create a button to show about window
about_button = tk.Button(root,text="about",command=show_about_window,width=100,height=2,background="#ABFFBC",font=("Arial", 12, "bold"))
about_button.pack(padx=5 ,pady=5)

# Create a log screen
log_frame = tk.Frame(root, padx=10, pady=10,width=100)
log_frame.pack(fill=tk.BOTH, expand=True)
log_frame.config(bg="#cceeff")
log_label = tk.Label(log_frame, text="Log Screen", anchor="w")
log_label.pack(fill=tk.X)

log_text = tk.Text(log_frame, bg="black", fg="white", state=tk.NORMAL,width=100)
log_text.pack(fill=tk.BOTH, expand=True)

# Redirect print statements to the log screen
sys.stdout = RedirectOutput(log_text)

# Start the Tkinter event loop
root.mainloop()
