import openpyxl
from openpyxl.styles import PatternFill
import random

from openpyxl import Workbook
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from datetime import datetime, timedelta
import re

import psutil
import subprocess
import winsound

red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
def is_chrome_running():
    # Check for running chrome instances
    for process in psutil.process_iter(['name']):
        if process.info['name'] and 'chrome' in process.info['name'].lower():
            return True
    return False

def prompt_user_to_close_chrome():
    # Prompt the user to close all chrome instances
    while is_chrome_running():
        input("Chrome is running. Please close all Chrome instances and press Enter to continue...")
    print("All Chrome instances are closed.")

def launch_chrome_in_debug_mode():
    # Launch Chrome in debugging mode
    chrome_path = 'C:/Program Files/Google/Chrome/Application/chrome.exe'  # Update this path if necessary
    subprocess.Popen([chrome_path, '--remote-debugging-port=9222', '--user-data-dir="C:\\users\\{os.getlogin()}\\AppData\\local\\Google\\chrome\\User Data"'])
    print("Chrome launched in debugging mode on port 9222.")
    
    
    
    

def create_excel(date, employee_data,cod_count_per_user,shipment_numbers, user_name):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Set the date in cell A1
    ws['A1'] = date

    # Write the header for the single user
    ws['B1'] = user_name

    
    # Write the data for the single user
    for row_num, value in enumerate(employee_data, start=2):
        ws.cell(row=row_num, column=2).value = value
        if value>10 :
            ws.cell(row=row_num, column=2).fill=red_fill
        
    for row_num  ,value in enumerate(shipment_numbers,start=2):
        ws.cell(row=row_num,column=3).value=value
        
    if len(employee_data) == 0 :
        employee_data=[1]
    ws.cell(row=len(employee_data)+3,column=1 ).value="Average = " 
    ws.cell(row=len(employee_data)+3,column=2 ).value= round(sum(employee_data)/len(employee_data),2)

    ws.cell(row=len(employee_data)+5,column=1 ).value="COD COUNT = " 
    ws.cell(row=len(employee_data)+5,column=2 ).value=len(cod_count_per_user)
    # Save the workbook
    file_name = f"{date.replace('/', '-')}_for_{user_name}.xlsx"
    wb.save(file_name)
    print(f"Excel file '{file_name}' created successfully.")
name=""
employees=[]
name_paths= {}
#dates={}
def get_employee_data_from_excel(input_path):

    #while not ((name:=input("Enter Name of Employee , Leave Empty to exit ")) == ""):
        #employees.append(name)
        #name_paths[name]=input(f"Enter Excel File Path for {name}")
        #while True:
            #dates[name]=input(f"Enter The Date of the File for {name} in the form mm-dd Ex. 05-27")
            #if not pattern.match(dates[name]): 
                #dates[name]=input(f"Enter The Date of the File for {name} in the form mm-dd Ex. 05-27")
            #else:
                #break 
    wb_input=openpyxl.load_workbook(input_path,data_only=True)
    ws_input=wb_input.active
        
    for i,row  in enumerate(ws_input): #skip first row (Headers)
        if i==0 :
            continue
        name =row[0].value
        path= row[1].value + '.xlsx'
        file_date =row[2].value
        is_random=row[3].value
        print (f'name : {name} , date : {file_date} , path : {path}')
        
        time_difference_per_user = [] 
        cod_count_per_user=[]
        cod_count=0
        shipment_numbers=[]
        # Open the provided Excel file and read the B column from row 2 onward in the first sheet
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        tracking_numbers = []
        for cell in ws['B'][1:]:  # Skip the first row (header)
            if cell.value:
                tracking_numbers.append(cell.value)

        if is_random=="1":
                print("Random of 20 Samples are Chosen")
                tracking_numbers= get_random_tracking_numbers(tracking_numbers) 
        else:
                print("Full File is Chosen")
        
        for number in tracking_numbers:
            print(f"{tracking_numbers.index(number)}: Working On {name} with Number : {number} in the Date : {file_date}\n")
            driver.execute_script(f"window.open('https://opost.ps/resources/shipments?tracking_number=gz-48-003711980', '_self');")

            # Switch to the new tab
            driver.switch_to.window(driver.window_handles[-1])

            # Optionally, wait for the page to fully load
            time.sleep(3)

            # Find the 29th button on the page and click it
            buttons = driver.find_elements(By.TAG_NAME, "button")
            if len(buttons) > 28:
                button = buttons[28]
                button.click()
            else:
                print("Button not found")
                winsound.Beep(600,1000) 
                continue

            # Wait for the new content to load
            try:
                # Define the locator for the new content
                new_content_locator = (By.XPATH, "//button[contains(., 'notes')]")  # the notes text is the only unique modifier  in the pop up opost dialog , it's text is never taken along the html flow of the page
                
                #this allocation is used only to "Take Time " , we give the system few time to make sure the content is fully loaded 
                new_content = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located(new_content_locator)
                )
                
                #The following allocation is the real content allocation , from the loaded content below we extract data
                new_content_locator = (By.CSS_SELECTOR, "tr")  # since the dialog is a table of rows , the main elemetn to target is the table rows collection
                new_content = WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located(new_content_locator)
                )
                
                # Once the new Table  is loaded, we can now parse the table or other new elements
                table_row = driver.find_elements(*new_content_locator)  #   retrieves  the collection of all rows 
                first_pending_of_driver =""                             #   define a variable to carry the first pending status 
                first_pending_of_employee=""                            #   define a variable to carry the first employee pending just after the driver firtst pending 
                
                
                is_first_time_employee_pen_detected=False                            #   since we catch all the dialog rows , we are to count the number of COD for each number in a defined date, 
                                                                        #   the COD status may repeate as well as the Pending status , we are targetting the first pending only 
                                                                        #   so the {is_first_time_detected} is put to control that logic
                is_first_time_driver_pen_detected=False
                # Table rows collection  is firstly reversed , to start from the 'Submitted'  status and forth on   
                # The logic below consists the results to the Pending and COD status types only, other status types are eventually ignored                                           
                # there is a bug here where the readign start from above
                for row in reversed(table_row):

                    
                    if  "Pending" in row.text and file_date in row.text:
                        td_elements = row.find_elements(By.CSS_SELECTOR, "td")

                        # Iterate through td elements and print their text
                        pending_data = [td.text for td in td_elements]
                        print(f"{pending_data[1]} {pending_data[3]}")
                        
                        if not is_first_time_driver_pen_detected:
                            if pending_data[1] == pending_data[3]:
                                first_pending_of_driver=pending_data[0]
                                is_first_time_driver_pen_detected=True
                            
                        if not is_first_time_employee_pen_detected :
                            if is_first_time_driver_pen_detected:
                                if '291لارا' in  pending_data[1]  or '296هبة' in  pending_data[1] or '290رند' in  pending_data[1] or '294حمزة' in  pending_data[1] or 'احمد295' in  pending_data[1] or 'متابعة عوالق' in  pending_data[1] :
                                    first_pending_of_employee=modify_time_if_before_10(pending_data[0])
                                    is_first_time_employee_pen_detected=True
                                #break
                    if "COD Pickup" in row.text and file_date in row.text:
                        td_elements = row.find_elements(By.CSS_SELECTOR, "td")

                        cod_date= [td.text for td in td_elements]
                        if file_date in cod_date[0]:
                            #winsound.Beep(500,500)
                            cod_count=cod_count+1
                            break
                        
                        
                        
                    pending_data=[]
                    

                        
                print(f"time of employee {first_pending_of_employee} \ntime of driver {first_pending_of_driver}\n") #   show Pending resluts for both employee and driver
                time1 = datetime.strptime(first_pending_of_driver, "%Y-%m-%d %H:%M:%S").time()                      #   extract the time only from the full date-time format of driver Pending status 
                time2 = datetime.strptime(first_pending_of_employee, "%Y-%m-%d %H:%M:%S").time()                    #   extract the time only from the full date-time format of employee Pending status 

                print(f"Time1 {time1}")
                print(f"Time2 {time2}")
                
                # Convert the time components to timedelta objects 
                time1_delta = timedelta(hours=time1.hour, minutes=time1.minute, seconds=time1.second)
                time2_delta = timedelta(hours=time2.hour, minutes=time2.minute, seconds=time2.second)

                print(f"Timedelta1 {time1_delta}")
                print(f"Timedelta2 {time2_delta}")
                # Calculate the absolute difference in minutes between the two time components
                time_difference = abs(time2_delta - time1_delta)
                difference_in_minutes = time_difference.total_seconds() / 60

                # Print the difference in minutes
                print(f"The difference in minutes is: {difference_in_minutes:.2f}")
                time_difference_per_user.append(round(difference_in_minutes,2))
                if cod_count > 0 :
                    cod_count_per_user.append(cod_count)
                shipment_numbers.append(number)
            
            
            except Exception as e:
                    print("New content did not load within the wait time:", e)
                    winsound.Beep(700,1000)
                    continue #  The flow should continue and ignore any exceptions as the exceptions are mainly generated pair Tracking Number 
                
        
        #call function to create the results as excel file 
        create_excel(file_date, time_difference_per_user,cod_count_per_user,shipment_numbers,name)
    
    
        
        
        

    
'''
def get_employee_data_from_user(employee_urls):
    employee_data = []

    for employee, url_entries in employee_urls.items():
        print(f"Enter 10 values for {employee}:")
        values = []
        for i in range(10):
            # Optionally, show the pre-selected random entry as a default value
            default_value = url_entries[i]
            value = input(f"Value {i + 1} (default {default_value}): ") or default_value
            values.append(value)
        employee_data.append(values)

    return employee_data
'''

def get_random_tracking_numbers(tracking_numbers_list):
    
    # Ensure the original list has at least 20 elements
    if len(tracking_numbers_list) >= 20:
        random_numbers = random.sample(tracking_numbers_list, 20)
        return random_numbers
    else:
        print("The list does not contain enough elements. Returning the whole tracking numbers list ")
        return  tracking_numbers_list

def modify_time_if_before_10(datetime_str):
    """
    Modify the time part of the datetime string to 10:00:00 if the hour is before 10:00:00.
    
    Parameters:
    datetime_str (str): The input datetime string in the format 'YYYY-MM-DD HH:MM:SS'.
    
    Returns:
    str: The modified datetime string.
    """
    # Parse the datetime string into a datetime object
    dt = datetime.strptime(datetime_str, '%Y-%m-%d %H:%M:%S')
    
    # Check if the hour is before 10
    if dt.hour < 10:
        # Modify the time to 10:00:00
        dt = dt.replace(hour=10, minute=0, second=0)
    
    # Convert the datetime object back to a string
    modified_datetime_str = dt.strftime('%Y-%m-%d %H:%M:%S')
    
    return modified_datetime_str



print('Starting Program .... ')

if is_chrome_running():
    prompt_user_to_close_chrome()
launch_chrome_in_debug_mode()


# Set up Chrome options to connect to the running instance
chrome_options = webdriver.ChromeOptions()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")

# Initialize the Chrome driver with the options
driver = webdriver.Chrome(options=chrome_options)

time_difference_per_user = [] 
shipment_numbers= []
pattern = re.compile(r'^(0[1-9]|1[0-2])-(0[1-9]|[12][0-9]|3[01])$')
#is_random="10"
#while  True :
#    is_random = input("If you want a random sample enter 1 \nif you want a full sample enter 2  ")  
#    if is_random in ["1","2"]:
#        break
    

# Get the date from the user
#date = input("Enter the date (YYYY/MM/DD): ")

# Get the tracking numbers from the Excel files and generate URLs
input_path= input ("Enter The Path of input Excel File that contains Names , Pathes and Dates\n")
employee_urls = get_employee_data_from_excel(input_path)

# Get the data for each employee from the user
#employee_data = get_employee_data_from_user(employee_urls)

driver.quit()
